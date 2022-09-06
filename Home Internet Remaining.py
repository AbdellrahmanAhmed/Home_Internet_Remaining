import ctypes
import os
import shutil
import sys
from datetime import date
from datetime import datetime, timedelta

from openpyxl import load_workbook
from openpyxl.styles import Alignment
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from winotify import Notification

os.chdir("C:/Users/Abdel/Desktop/Home Internet Remaining")

ctypes.windll.user32.ShowWindow(ctypes.windll.kernel32.GetConsoleWindow(), 6)


# class WindowsBalloonTip:
#     def __init__(self, title, msg):
#         message_map = {
#             win32con.WM_DESTROY: self.OnDestroy,
#         }
#         # Register the Window class.
#         wc = WNDCLASS()
#         hinst = wc.hInstance = GetModuleHandle(None)
#         wc.lpszClassName = "PythonTaskbar"
#         wc.lpfnWndProc = message_map  # could also specify a wndproc.
#         classAtom = RegisterClass(wc)
#         # Create the Window.
#         style = win32con.WS_OVERLAPPED | win32con.WS_SYSMENU
#         self.hwnd = CreateWindow(classAtom, "Taskbar", style,
#                                  0, 0, win32con.CW_USEDEFAULT, win32con.CW_USEDEFAULT,
#                                  0, 0, hinst, None)
#         UpdateWindow(self.hwnd)
#         iconPathName = os.path.abspath(os.path.join(sys.path[0], "balloontip.ico"))
#         icon_flags = win32con.LR_LOADFROMFILE | win32con.LR_DEFAULTSIZE
#         try:
#             hicon = LoadImage(hinst, iconPathName,
#                               win32con.IMAGE_ICON, 0, 0, icon_flags)
#         except:
#             hicon = LoadIcon(0, win32con.IDI_APPLICATION)
#         flags = NIF_ICON | NIF_MESSAGE | NIF_TIP
#         nid = (self.hwnd, 0, flags, win32con.WM_USER + 20, hicon, "tooltip")
#         Shell_NotifyIcon(NIM_ADD, nid)
#         Shell_NotifyIcon(NIM_MODIFY,
#                          (self.hwnd, 0, NIF_INFO, win32con.WM_USER + 20,
#                           hicon, "Balloon  tooltip", title, 200, msg))
#         # self.show_balloon(title, msg)
#         time.sleep(10)
#         DestroyWindow(self.hwnd)
#
#     def OnDestroy(self, hwnd, msg, wparam, lparam):
#         nid = (self.hwnd, 0)
#         Shell_NotifyIcon(NIM_DELETE, nid)
#         PostQuitMessage(0)  # Terminate the app.


# def balloon_tip(title, msg):
#     w = WindowsBalloonTip(msg, title)


# window = tk.Tk()
# window.title('Home Internet Remaining.xlsx')
# window.geometry('200x200')


# def some_job02():
#     wb = load_workbook('Home Internet Remaining.xlsx')
#
#     cell = 1
#
#     today = date.today()
#     now = datetime.now()
#
#     month = today.strftime("%B")
#     timeN = now.strftime("%H:%M:%S")
#     todayN = today.strftime("%A")
#
#     def writeCell(cell0):
#         ws[f'C${cell0}'].value = todayN
#         ws[f'B${cell0}'].value = month
#         ws[f'A${cell0}'].value = today
#         ws[f'D${cell0}'].value = timeN
#         ws[f'E${cell0}'].value = remaining
#
#     ws = wb.active
#     cellValue = ws[f'A${cell}'].value
#
#     remaining = float(input("Remaining\n"))
#
#     # print(remaining)
#
#     while cellValue:
#         cell = cell + 1
#         cellValue = ws[f'A${cell}'].value
#
#     writeCell(cell)
#
#     cellList = ['A', 'B', 'C', 'D', 'E']
#
#     for cellC in cellList:
#         currentCell = ws[f'${cellC}${cell}']
#         currentCell.alignment = Alignment(horizontal='center')
#
#     # print(cell)
#
#     lastRemaining = ws[f'E${cell - 1}'].value
#     (used) = round(float(lastRemaining) - float(remaining), 2)
#     # print(f'Last Used {round(used, 2)} GB')
#
#     last_Remaining = ws[f'E${cell - 1}'].value
#
#     tt = str(ws[f'D${cell - 1}'].value)
#     after_hour_date_time = datetime.now() + timedelta(hours=1)
#
#     bTime = datetime.strptime(tt, "%H:%M:%S")
#     btTime = datetime.strptime(timeN, "%H:%M:%S") - bTime
#
#     print(f'''
#             Time Now {timeN}        Remaining {remaining} GB
#
#             Last time Reading  {tt}  Last Remaining: {last_Remaining}
#
#             Used {used} GB                 Duration  {btTime}
#
#
#             ''')
#
#     # Next Reading {after_hour_date_time.strftime('%H:%M:%S')}
#
#     wb.save('Home Internet Remaining.xlsx')
#     time.sleep(2)


def some_job():
    global msg
    try:
        print("Decorated job")
        wb = load_workbook('Home Internet Remaining.xlsx')

        cell = 1

        today = date.today()
        now = datetime.now()

        month = today.strftime("%B")
        timeN = now.strftime("%H:%M:%S")
        todayN = today.strftime("%A")

        def writeCell(cell0):
            ws[f'A${cell0}'].value = today
            ws[f'B${cell0}'].value = month
            ws[f'C${cell0}'].value = todayN
            ws[f'D${cell0}'].value = timeN
            ws[f'E${cell0}'].value = remaining

        ws = wb.active
        cellValue = ws[f'A${cell}'].value

        chrome_options = Options()
        chrome_options.add_argument(
            "--user-data-dir=C:/Users/Abdel/AppData/Local/Google/Chrome/User Data/unique", )  # change to profile path
        chrome_options.add_argument('--profile-directory=Default')
        driver = webdriver.Chrome("C:\Program Files\Google\Chrome\Application\chromedriver.exe")

        driver.minimize_window()

        driver.get('https://my.te.eg/user/login')

        driver.implicitly_wait(3)

        service_Number = driver.find_element(by=By.XPATH, value='//*[@id="login-service-number-et"]')
        password = driver.find_element(by=By.XPATH, value='//*[@id="login-password-et"]')
        p_Button_Label = driver.find_element(by=By.XPATH, value='//*[@id="login-login-btn"]/span[2]')

        service_Number.send_keys('0663696485')
        password.send_keys('01552337484we')
        p_Button_Label.click()

        driver.implicitly_wait(5)

        # time.sleep(5)

        old_Remaining = driver.find_element(by=By.XPATH,
                                            value='//*[@id="pr_id_7"]/div/div/div/div/div/app-gauge/div[2]/span[1]').text

        # print(remaining0)
        remaining = ""
        for c in old_Remaining:
            if c == ".":
                remaining = remaining + c
            elif c.isdigit():
                remaining = remaining + c

        # print(remaining)

        while cellValue:
            cell = cell + 1
            cellValue = ws[f'A${cell}'].value

        writeCell(cell)

        cellList = ['A', 'B', 'C', 'D', 'E']

        for cellC in cellList:
            currentCell = ws[f'${cellC}${cell}']
            currentCell.alignment = Alignment(horizontal='center')

        # print(cell)

        lastRemaining = ws[f'E${cell - 1}'].value
        (used) = round(float(lastRemaining) - float(remaining), 2)
        # print(f'Last Used {round(used, 2)} GB')

        last_Remaining = ws[f'E${cell - 1}'].value
        remaining_days = date(2022, 9, 18) - today
        recommended_Remaining = round(float(remaining) / float(remaining_days.days), 2)

        tt = str(ws[f'D${cell - 1}'].value)
        after_hour_date_time = datetime.now() + timedelta(hours=1)

        bTime = datetime.strptime(tt, "%H:%M:%S")
        btTime = datetime.strptime(timeN, "%H:%M:%S") - bTime

        cT = []
        tSTR = f"{today} 00:00:00"
        for c in range(1, 1000):
            if str(ws[f'A${c}'].value) == str(tSTR):
                v = ws[f'E{c}'].value
                cT.append(float(v))
            elif not ws[f'A{c}'].value:
                cT.append(0.0)
                break
        totalDayUsed = float(cT[0]) - float(cT[-1])

        msg = f'''
        Remaining {remaining} GB 
        recommended consumption  {recommended_Remaining} GB 
        and Total Day Remaining {round(totalDayUsed, 2)}GB
        '''
        print(msg)

        # Next Reading {after_hour_date_time.strftime('%H:%M:%S')}

        driver.close()
        wb.save('Home Internet Remaining.xlsx')
        # time.sleep(2)
        # lblR['text'] = remaining
        # lblU['text'] = used
        # lblD['text'] = btTime
        # print("Next Time Running:  ")
        # nextTime =datetime.today()+timedelta(minutes=Minutes)
        # print(nextTime)

        shutil.copy(f'Home Internet Remaining.xlsx', f'Home Internet Remaining {today}.xlsx')
        # WindowsBalloonTip.balloon_tip(f"Done..",f"Remaining {remaining} GB")
        # sys.exit()

        return str(msg)


    except Exception as e:
        # lblD['text'] = 'Try Again'
        print('***************************************************************')
        print(e)
        print('***************************************************************')

        # inputTry = None
        #
        # def check():
        #     time.sleep(2)
        #     if inputTry is None:
        #         print("Too Slow")
        #     if inputTry == '1':
        #         driver.close()
        #         some_job()
        #     elif inputTry == '2':
        #         driver.close()
        #         some_job02()
        #         # print("Next Time Running:  ")
        #         # nextTime = datetime.today() + timedelta(minutes=Minutes)
        #         # print(nextTime)
        #     elif inputTry is None:
        #         driver.close()
        #         some_job()
        #     else:
        #         sys.exit()
        #
        #
        #
        # Thread(target=check).start()
        #
        # inputTry = input('''
        # [1]for Again
        # [2]Manual
        # [3]Close \n
        #                  ''')

        driver.close()
        some_job()

    return str(msg)


#
# btn = tk.Button(text='calculate', command=some_job())
# btn.place(x=50, y=50)
#
# lblR = tk.Label(window, text="", font=("Helvetica", 16))
# lblU = tk.Label(window, text="", font=("Helvetica", 16))
# lblD = tk.Label(window, text="", font=("Helvetica", 16))

# btn.pack()
# lblR.pack()
# lblU.pack()
# lblD.pack()
# window.mainloop()

# Minutes = input('minutes = ')
# some_job()
# scheduler = BlockingScheduler()
# scheduler.add_job(some_job, 'interval', minutes=int(Minutes))
# scheduler.start()

# balloon_tip("sasds", "sasas")

# tRemaining = some_job


toast = Notification(
    app_id="Home Internet",
    title="تم حساب المتبقي من الإنترنت وتخزينه",
    msg=some_job(),
    duration="long",
)

toast.add_actions(label="للدخول إلى صفحتك الشخصية  ", launch="https://my.te.eg/user/login")
# openExcelFile()

toast.show()

sys.exit()
